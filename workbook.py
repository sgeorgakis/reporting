#!/usr/bin/env

import openpyxl
import datetime
from openpyxl.chart import (
    LineChart,
    Reference,
    Series
)
import pdb
import re

from openpyxl.chart.axis import DateAxis

raw_data_sheet = 'raw'
chart_sheet = 'Charts'
data_sheet = 'Data'
raw_chart_data_sheet = 'raw_chart_data'
chart_spend_title = 'Daily Spend Uplift (ALL Brands)'
chart_topup_title = 'Daily TopUp Uplift (ALL Brands)'
chart_spend_legend = 'Spend Lift'
chart_topup_legend = 'TopUp Lift'
chart_date_legend = 'Date'


class Workbook:

    def __init__(self, workbook_name, index, offset):
        self.__wb = openpyxl.load_workbook(
            workbook_name
            )
        self.__working_sheet = None
        self.__index = index
        self.__offset = offset
        name = re.split('\d{4}_\d{2}_\d{2}', workbook_name)
        date = datetime.datetime.now().strftime("%Y_%m_%d")
        self.__new_file_name = date + name[1]

    def open_sheet(self, sheet_name):
        self.__working_sheet = self.__wb.get_sheet_by_name(
            sheet_name
        )

    def __add_new_data_line(self, data):
        title = self.__working_sheet.title
        sheet = self.__wb.create_sheet(
            title=(self.__working_sheet.title + '.1')
        )
        sheet.sheet_properties = self.__working_sheet.sheet_properties

        # Copy header if any
        for row_of_cell_objects in self.__working_sheet['A1':'R1']:
            for cell_obj in row_of_cell_objects:
                sheet[cell_obj.coordinate] = cell_obj.value

        # Create a new row after the header
        saved_format = False
        # skip the first row if needed
        for row in range(2, self.__working_sheet.max_row):
            for column in range(1, self.__working_sheet.max_column):
                sheet.cell(
                    row=row+1,
                    column=column
                    ).value = self.__working_sheet.cell(
                        row=row,
                        column=column
                        ).value
                # Copy also format of the cell.
                # It is mostly useful for the date cells
                if self.__working_sheet.cell(
                    row=row,
                    column=column
                ).has_style:
                    sheet.cell(
                        row=row+1,
                        column=column
                        ).font = self.__working_sheet.cell(
                            row=row,
                            column=column
                            ).font
                    sheet.cell(
                        row=row+1,
                        column=column
                        ).border = self.__working_sheet.cell(
                            row=row,
                            column=column
                            ).border
                    sheet.cell(
                        row=row+1,
                        column=column
                        ).fill = self.__working_sheet.cell(
                            row=row,
                            column=column
                            ).fill
                    sheet.cell(
                        row=row+1,
                        column=column
                        ).number_format = self.__working_sheet.cell(
                            row=row,
                            column=column
                            ).number_format
                    sheet.cell(
                        row=row+1,
                        column=column
                        ).protection = self.__working_sheet.cell(
                            row=row,
                            column=column
                            ).protection
                    sheet.cell(
                        row=row+1,
                        column=column
                        ).alignment = self.__working_sheet.cell(
                            row=row,
                            column=column
                            ).alignment
                    if ((not saved_format) and
                        (isinstance(self.__working_sheet.cell(
                            row=row,
                            column=column
                            ).value, datetime.datetime))):
                        datetime_format = self.__working_sheet.cell(
                            row=row,
                            column=column
                            )
                        saved_format = True

        # Insert data
        column_index = 1
        for cell in data:
            sheet.cell(row=2, column=column_index).value = cell
            if column_index == 1 or column_index == 2:
                sheet.cell(
                    row=2,
                    column=column_index
                    ).font = datetime_format.font
                sheet.cell(
                    row=2,
                    column=column_index
                    ).border = datetime_format.border
                sheet.cell(
                    row=2,
                    column=column_index
                    ).fill = datetime_format.fill
                sheet.cell(
                    row=2,
                    column=column_index
                    ).number_format = datetime_format.number_format
                sheet.cell(
                    row=2,
                    column=column_index
                    ).protection = datetime_format.protection
                sheet.cell(
                    row=2,
                    column=column_index
                    ).alignment = datetime_format.alignment
            column_index = column_index + 1

        # Remove old raw sheet
        self.__wb.remove_sheet(self.__working_sheet)
        self.__working_sheet = sheet
        self.__working_sheet.title = title

        self.__datetime_format = datetime_format

    def add_data(self, data):
        for data_line in data:
            self.__add_new_data_line(data_line)
        self.__wb.save(self.__new_file_name)

    def create_line_chart(self):
        c1 = LineChart()
        c1.title = chart_spend_title
        c1.style = 12
        c1.y_axis.title = "Euros"
        c1.y_axis.crossAx = 500
        c1.x_axis = DateAxis(crossAx=100)
        c1.x_axis.number_format = 'd-mm-yy'
        c1.x_axis.majorTimeUnit = "days"
        self.open_sheet(raw_chart_data_sheet)
        values = Reference(
            self.__working_sheet,
            min_row=1,
            max_row=self.__working_sheet.max_row,
            min_col=3,
            max_col=3
            )
        series = Series(values, title_from_data=True)
        c1.append(series)
        dates = Reference(
            self.__working_sheet,
            min_col=1,
            max_col=1,
            min_row=1,
            max_row=self.__working_sheet.max_row)
        c1.set_categories(dates)
        c1.height = 30
        c1.width = 50

        c2 = LineChart()
        c2.title = chart_topup_title
        c2.style = 12
        c2.y_axis.title = "Euros"
        c2.y_axis.crossAx = 500
        c2.x_axis = DateAxis(crossAx=100)
        c2.x_axis.number_category = 'Date'
        c2.x_axis.number_format = 'd-mm-yy'
        c2.x_axis.majorTimeUnit = "days"
        self.open_sheet(raw_chart_data_sheet)
        values = Reference(
            self.__working_sheet,
            min_row=1,
            max_row=self.__working_sheet.max_row,
            min_col=2,
            max_col=2
            )
        series = Series(values, title_from_data=True)
        c2.append(series)
        dates = Reference(
            self.__working_sheet,
            min_col=1,
            max_col=1,
            min_row=1,
            max_row=self.__working_sheet.max_row)
        c2.set_categories(dates)
        c2.height = 30
        c2.width = 50

        self.open_sheet(chart_sheet)
        self.__working_sheet.add_chart(c1, "C1")
        self.__working_sheet.add_chart(c2, "C60")

        self.__wb.save(self.__new_file_name)
        self.open_sheet(raw_data_sheet)

    def update_raw_data_chart(self):
        self.open_sheet(data_sheet)
        data = []

        # today = datetime.datetime.today().day - 1
        today = 12
        for row in range(
                         self.__offset,
                         today * self.__index + self.__offset,
                         self.__index
                        ):
            inner_data = []
            inner_data.append(
                self.__working_sheet.cell(
                    row=row,
                    column=1).internal_value
                )
            inner_data.append(
                self.__working_sheet.cell(
                    row=row,
                    column=3).internal_value
                )
            inner_data.append(
                self.__working_sheet.cell(
                    row=row,
                    column=5).internal_value
                )
            data.append(inner_data)
        self.open_sheet(raw_chart_data_sheet)

        data.reverse()
        self.__add_chart_data(data)
        self.__wb.save(self.__new_file_name)

    def __add_chart_data(self, data):
        # Add headers
        self.__working_sheet.cell(
            row=1,
            column=1
            ).value = chart_date_legend
        self.__working_sheet.cell(
            row=1,
            column=2
            ).value = chart_topup_legend
        self.__working_sheet.cell(
            row=1,
            column=3
            ).value = chart_spend_legend
        row = 2
        for inner_data in data:
            self.__working_sheet.cell(
                row=row,
                column=1
                ).value = inner_data[0]
            self.__working_sheet.cell(
                row=row,
                column=1
                    ).number_format = self.__datetime_format.number_format

            self.__working_sheet.cell(
                row=row,
                column=2
                ).value = inner_data[1]
            self.__working_sheet.cell(
                row=row,
                column=3
                ).value = inner_data[2]
            row += 1

if __name__ == "__main__":

    workbook = Workbook(
       "2016_09_12_DailyReportHybrid.xlsx",
       1,
       4
       )

    workbook.open_sheet(raw_data_sheet)
    all_data = ["",
                [
                    '1-Jul-16', '21-Jul-16',
                    '22-JUL-16 01.47.38.362000000 PM',
                    'ALL', 335224, 3886925, 1, 672420.849047997,
                    7909382.96088397, 83835, 983746,
                    789202.412, 9285849.941, 1242794, 105979,
                    2915922, 2915922, 130246
                ],
                [
                    '1-Jul-16', '21-Jul-16',
                    '22-JUL-16 01.47.38.362000000 PM',
                    'Unknown', 956, 10855, 1, 116.928173, 891.852316,
                    6, 74, 75, 811, 107, 10, 202, 202, 17
                ],
                [
                    '1-Jul-16', '21-Jul-16',
                    '22-JUL-16 01.47.38.362000000 PM',
                    'TAZA', 29850, 343980, 0, 65392.731086,
                    753939.545907, 10043, 116532, 73057.401, 850568.641,
                    81831, 7140, 232634, 232634, 17370
                ],
                [
                    '1-Jul-16', '21-Jul-16',
                    '22-JUL-16 01.47.38.362000000 PM',
                    'OCFP', 1499, 17829, 1, 3187.86597, 41938.7604629999,
                    344, 4001, 3512, 41333, 5709, 467, 13503, 13503, 548
                ],
                [
                    '1-Jul-16', '21-Jul-16',
                    '22-JUL-16 01.47.38.362000000 PM',
                    'New VF Prepay', 73179, 855901, 1, 150852.330815999,
                    1814572.95918399, 17950, 214516, 179298.6, 2148076.6,
                    313665, 26311, 699340, 699340, 38489
                ],
                [
                    '1-Jul-16', '21-Jul-16',
                    '22-JUL-16 01.47.38.362000000 PM',
                    'International', 64853, 754366, 1, 89972.537244,
                    1050336.44697199, 10434, 120625, 103612, 1207438.8,
                    129122, 10829, 640221, 640221, 14463
                ],
                [
                    '1-Jul-16', '21-Jul-16',
                    '22-JUL-16 01.47.38.362000000 PM',
                    'CU', 164632, 1901413, 1, 362662.963839998,
                    4242592.39606297, 45038, 527552, 429438.411,
                    5032786.9, 711506, 61153, 1329901, 1329901, 59357
                ]
                ]
    hybrid_data = [
                '1-Sep-16', '12-Sep-16',
                '15-SEP -16 12.28.56,053000000 PM',
                'Kartoprogramma', 15397, 176510, 0, 4609.999731,
                52609.46223, 471, 4809, 4976, 49719, 27821,
                2355, 126743, 126743, 2716
                  ]
    workbook.add_data([hybrid_data])
    # workbook.add_data(all_data)
    workbook.update_raw_data_chart()
    workbook.create_line_chart()
